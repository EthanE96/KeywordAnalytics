import os
import glob
import sys
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Color


# Path to the folder containing the Excel files
folder_path = '/Users/ethan/Library/CloudStorage/OneDrive-Personal/Ecomm/EddysCove/Products/Research/Erank Data'


# Get the most recent Excel file from the folder
excel_files = glob.glob(os.path.join(folder_path, '*.xlsx'))
most_recent_file = max(excel_files, key=os.path.getctime)

# Extract the base name from the file path
file_name = os.path.basename(most_recent_file)

# Remove the extension from the file name
file_name_without_extension = os.path.splitext(file_name)[0]

# Load the most recent Excel file
wb = load_workbook(most_recent_file)

# Get the first sheet of the workbook
ws = wb.active

# Check if complete
if file_name_without_extension in wb.sheetnames:
    sys.exit("Sheet named: " + file_name_without_extension +
             " already exists. Script ended. Workbook: " + most_recent_file)

# Create a new sheet within the same workbook
ns = wb.create_sheet(title=file_name_without_extension)

# Copy the data from the source sheet to the new sheet
for row in ws.iter_rows(values_only=True):
    ns.append(row)

# Get the range of cells containing the copied data
num_rows = ns.max_row
num_columns = ns.max_column
start_cell = ns.cell(row=1, column=1)
end_cell = ns.cell(row=num_rows, column=num_columns)

# Create a table from the range of cells
table = Table(displayName="Table1",
              ref=f"{start_cell.coordinate}:{end_cell.coordinate}")

# Apply table style
table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium8")

# Add the table to the new sheet
ns.add_table(table)

# Delete columns 1 and 2
ns.delete_cols(idx=1, amount=2)

# Insert new columns
ns.insert_cols(idx=9, amount=2)
column_name_1 = "Comp to Search"
column_name_2 = "Search Weighted"
column_letter_1 = get_column_letter(9)
column_letter_2 = get_column_letter(10)
ns[f"{column_letter_1}1"] = column_name_1
ns[f"{column_letter_2}1"] = column_name_2


# Fill column Comp to Search with this formula (F2/C2)
for row in range(2, num_rows + 1):
    cell = ns[f"I{row}"]
    cell.value = f"=F{row}/C{row}"

# Fill column J with this formula (I2/C2)
for row in range(2, num_rows + 1):
    cell = ns[f"J{row}"]
    cell.value = f"=I{row}/C{row}"
    cell.number_format = "0.00"

# Define the minimum, midpoint, and maximum values
minimum_value = 0
midpoint_value = 5
maximum_value = 10

# Get the range of cells to apply the color scale rule
num_rows = ns.max_row
start_cell = ns.cell(row=2, column=9)  # Cell I2
end_cell = ns.cell(row=num_rows, column=9)  # Last cell in column I
column_letter = get_column_letter(9)  # Column I letter
range_string = f"{column_letter}{start_cell.row}:{column_letter}{end_cell.row}"

# Create a color scale rule
color_scale_rule = ColorScaleRule(
    start_type='num',
    start_value=minimum_value,
    start_color=Color(rgb='63BE7B'),  # Green color
    mid_type='num',
    mid_value=midpoint_value,
    mid_color=Color(rgb='FFEB84'),  # Yellow color
    end_type='num',
    end_value=maximum_value,
    end_color=Color(rgb='F8696B')  # Red color
)

# Apply the color scale rule to the range of cells
ns.conditional_formatting.add(range_string, color_scale_rule)

# Format the column width of the table
for column in ns.columns:
    column_letter = column[0].column_letter
    ns.column_dimensions[column_letter].width = 12

# Save the copied workbook
wb.save(most_recent_file)

print("Sheet named " + file_name_without_extension + " complete")
